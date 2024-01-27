import json
import re
import openpyxl

def process_excel_to_json(input_file,symptom_file, output_file, output_file2):
    rows = []
    rows2 = []
    count=0
    new_wb_symptom = openpyxl.Workbook()
    new_ws_symptom = new_wb_symptom.active
    new_ws_symptom['A1'] = 'name'
    new_ws_symptom['B1'] = 'symptom'
    new_ws_symptom['C1'] = 'method'
    
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws['A1'] = 'name'
    new_ws['B1'] = 'part'
    new_ws['C1'] = 'alias'
    new_ws['D1'] = 'smell'
    new_ws['E1'] = 'cure'
    efficacy = dict()
    symptom = dict()

    symwb = openpyxl.load_workbook(symptom_file)
    symsheet = symwb["Sheet1"]
    for row in symsheet.iter_rows(min_row=2, max_col=4, values_only=True):
        if row[1] == '功效':
            if row[0] not in efficacy.keys():
                efficacy[row[0]] = (row[2] if row[2][-1]=='。' else (row[2]+"。"))
            else:
                efficacy[row[0]] += (row[2] if row[2][-1]=='。' else (row[2]+"。"))
        else:
            if row[0] not in symptom.keys():
                #print(row)
                symptom[row[0]] = {row[2]:(row[3] if row[3][-1]=='。' else (row[3]+"。"))}
            elif row[2] not in symptom[row[0]].keys() :
                symptom[row[0]][row[2]] = (row[3] if row[3][-1]=='。' else (row[3]+"。"))
            else:
                symptom[row[0]][row[2]] += (row[3] if row[3][-1]=='。' else (row[3]+"。"))
            
    for k in symptom.keys():
        for s in symptom[k].keys():
            rows2.append([k,s,symptom[k][s]])
    
    # Load the workbook
    
    wb = openpyxl.load_workbook(input_file)
    sheet = wb["Sheet1"]

    # Iterate through each row in column A and D
    for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True):
        print(row[1])
        cure1 = ''
        cure2 = ''
        if row[1] in efficacy.keys():
            cure1 = efficacy[row[1]]
        if row[1] in symptom.keys():
            cure2 = "主治："+';'.join(symptom[row[1]].keys())
        if len(cure1+cure2)>0:
            new_row=[row[1],row[0],row[2],row[3],cure1+cure2]
        else:
            new_row=[row[1],row[0],row[2],row[3]]
        rows.append(new_row)

    for new_row2 in rows2:
        new_ws_symptom.append(new_row2)
    for new_row in rows:
        new_ws.append(new_row)
        count+=1
    print(count)
    new_wb.save(output_file)
    new_wb_symptom.save(output_file2)


process_excel_to_json('./data/xlsx_all/ChineseMedicalData.xlsx','./symptom.xlsx', './data/xlsx_new/ChineseMedicalNew.xlsx', './data/xlsx_new/SymptomNew.xlsx')
