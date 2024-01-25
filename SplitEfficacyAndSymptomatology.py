import json
import re
import openpyxl

def process_excel_to_json(input_files, output_file):
    rows = []
    rows=[]
    count=0
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws['A1'] = 'name'
    new_ws['B1'] = 'Symptom'
    new_ws['C1'] = 'method'
    for input_file in input_files:
        # Load the workbook
        
        wb = openpyxl.load_workbook(input_file)
        sheet = wb["Sheet1"]

        # Iterate through each row in column A and D
        for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True):
            
            #a="1、胸痹急痛（痛如锥刺，不能俯仰，自汗）。用生韭或根五斤，洗净捣汁服。2、阴阳易病（男子因房事不慎，引起阴部肿大，小腹绞痛，头重眼花）。用鼠尿十四粒、韭根一大把，同煮开，去渣，再煮开二次，温服，得汗即愈，无汗可再服一剂。3、伤寒劳复（按：指伤寒病后，身体未复原而性交，引起旧病复发）。治方同上。4、喘息欲绝。取韭汁一升饮下。5、盗汗。用韭根四十九根。加水二升煮成一升，一次服下。6、消渴。有徘苗或炒或作汤。日食三、五两，可加酱，但不可加盐。吃至十斤即见效。过了清明节，不宜此方。7、痢疾。多吃韭菜，作汤，煮粥，炒吃都行。8、赤白带下。用韭根捣汁，加童便露一夜，空心温服。9、疮癣。用大韭根炒存性。捣为末。调猪油涂搽。10、刀伤出血。用韭汁拌风化石灰，晒干，研为末，敷疮上。11、漆疮作痒。用韭叶捣烂敷上。12、耳出汁。用韭汁滴耳中，一天三次。13、食物中毒。用生韭汁数升可解。"
            cure = row[4]
            b=re.sub("。([0-9]+、)","。|\\1",cure)

            for i in b.split("|"):
                symptom,method = i[:i.find('。')],i[i.find('。')+1:]
                symptom = re.sub('[0-9]+、','',symptom)
                print(symptom,method)

                new_row=[row[1],symptom,method]

                rows.append(new_row)
    for new_row in rows:
        new_ws.append(new_row)
        count+=1
    print(count)
    new_wb.save(output_file)


process_excel_to_json(['./data/xlsx_all/ChineseMedicalData.xlsx'], './output.xlsx')
